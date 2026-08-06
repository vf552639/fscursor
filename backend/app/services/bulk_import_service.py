import csv
import io
import uuid
from dataclasses import dataclass
from typing import Optional
from uuid import UUID

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.validators import is_valid_domain, normalize_domain
from app.schemas.domain import DomainBulkCreateItem, DomainBulkImportError
from app.schemas.server import ServerBulkImportError
from app.services import domain_service
from app.services import server_service
from app.schemas.server import ServerCreate

_ERROR_EXPORTS: dict[str, str] = {}


def _parse_csv(raw: bytes, has_header: bool) -> list[tuple[int, str, Optional[str]]]:
    text = raw.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows: list[tuple[int, str, Optional[str]]] = []
    for idx, row in enumerate(reader, start=1):
        if has_header and idx == 1:
            continue
        if not row:
            continue
        domain = (row[0] or "").strip()
        registrar_name = (row[1] or "").strip() if len(row) > 1 else None
        rows.append((idx, domain, registrar_name or None))
    return rows


def _parse_xlsx(raw: bytes, has_header: bool) -> list[tuple[int, str, Optional[str]]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows: list[tuple[int, str, Optional[str]]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if has_header and idx == 1:
            continue
        if not row:
            continue
        domain = str(row[0] or "").strip()
        registrar_name = str(row[1] or "").strip() if len(row) > 1 and row[1] else None
        if not domain:
            continue
        rows.append((idx, domain, registrar_name))
    return rows


def build_errors_csv(errors: list[DomainBulkImportError]) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["row", "domain", "reason"])
    for e in errors:
        w.writerow([e.row, e.domain, e.reason])
    return out.getvalue()


def store_errors_csv(csv_text: str) -> str:
    token = uuid.uuid4().hex
    _ERROR_EXPORTS[token] = csv_text
    return token


def get_errors_csv(token: str) -> Optional[str]:
    return _ERROR_EXPORTS.get(token)


@dataclass
class ServerImportRow:
    row: int
    name: str
    ip: str
    ssh_user: str
    ssh_password: str
    ssh_port: int
    # Шестая, опциональная колонка — в конце, чтобы файлы без неё (старый
    # формат) продолжали импортироваться без изменений.
    provider: Optional[str] = None


async def process_bulk_import(
    db: AsyncSession,
    *,
    user_id: UUID,
    filename: str,
    content: bytes,
    has_header: bool,
    default_registrar_id: Optional[int] = None,
) -> tuple[int, int, list[DomainBulkImportError], Optional[str]]:
    name = filename.lower()
    if name.endswith(".xlsx"):
        rows = _parse_xlsx(content, has_header)
    else:
        rows = _parse_csv(content, has_header)

    valid_items: list[DomainBulkCreateItem] = []
    errors: list[DomainBulkImportError] = []
    for row_num, domain, registrar_name in rows:
        norm = normalize_domain(domain)
        if not is_valid_domain(norm):
            errors.append(
                DomainBulkImportError(row=row_num, domain=domain, reason="Invalid domain format")
            )
            continue
        valid_items.append(
            DomainBulkCreateItem(
                domain_name=norm,
                registrar_id=default_registrar_id,
                registrar_name=registrar_name,
            )
        )

    created_list, skipped = await domain_service.bulk_create_structured(
        db, user_id, valid_items
    )
    skipped_set = set(skipped)
    for row_num, domain, _ in rows:
        norm = normalize_domain(domain)
        if norm in skipped_set:
            errors.append(
                DomainBulkImportError(row=row_num, domain=domain, reason="Duplicate or exists")
            )

    csv_url = None
    if errors:
        token = store_errors_csv(build_errors_csv(errors))
        csv_url = f"/domains/bulk-import-errors/{token}"

    return len(created_list), len(skipped), errors, csv_url


def _parse_server_row(row: list[str], idx: int) -> Optional[ServerImportRow]:
    if not row:
        return None
    name = str(row[0] or "").strip()
    ip = str(row[1] or "").strip() if len(row) > 1 else ""
    ssh_user = str(row[2] or "").strip() if len(row) > 2 else "root"
    ssh_password = str(row[3] or "").strip() if len(row) > 3 else ""
    raw_port = str(row[4] or "").strip() if len(row) > 4 else "22"
    try:
        ssh_port = int(raw_port or "22")
    except ValueError:
        ssh_port = -1
    # Провайдер — свободный текст, набранный человеком в файле, а не
    # write-back с машины: обрезаем пробелы здесь же, как и остальные
    # текстовые поля строки. Саму валидацию (длина, управляющие символы)
    # намеренно не дублируем — этим занимается `_checked_provider` на
    # `ServerCreate`, а невалидное значение уходит в errors штатным путём.
    provider_raw = str(row[5] or "").strip() if len(row) > 5 else ""
    provider = provider_raw or None
    return ServerImportRow(
        row=idx,
        name=name,
        ip=ip,
        ssh_user=ssh_user or "root",
        ssh_password=ssh_password,
        ssh_port=ssh_port,
        provider=provider,
    )


def _parse_servers_csv(raw: bytes, has_header: bool) -> list[ServerImportRow]:
    text = raw.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows: list[ServerImportRow] = []
    for idx, row in enumerate(reader, start=1):
        if has_header and idx == 1:
            continue
        parsed = _parse_server_row(row, idx)
        if parsed:
            rows.append(parsed)
    return rows


def _parse_servers_xlsx(raw: bytes, has_header: bool) -> list[ServerImportRow]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows: list[ServerImportRow] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if has_header and idx == 1:
            continue
        normalized = ["" if cell is None else str(cell) for cell in row]
        parsed = _parse_server_row(normalized, idx)
        if parsed:
            rows.append(parsed)
    return rows


def build_server_errors_csv(errors: list[ServerBulkImportError]) -> str:
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["row", "server", "reason"])
    for err in errors:
        writer.writerow([err.row, err.server, err.reason])
    return out.getvalue()


async def process_server_bulk_import(
    db: AsyncSession,
    *,
    user_id: UUID,
    filename: str,
    content: bytes,
    has_header: bool,
) -> tuple[int, int, list[ServerBulkImportError], Optional[str]]:
    name = filename.lower()
    rows = _parse_servers_xlsx(content, has_header) if name.endswith(".xlsx") else _parse_servers_csv(content, has_header)

    existing = await server_service.get_all(db, user_id)
    existing_ips = {srv.ip_address for srv in existing[0]}

    created = 0
    skipped = 0
    errors: list[ServerBulkImportError] = []

    for item in rows:
        if not item.name:
            skipped += 1
            errors.append(ServerBulkImportError(row=item.row, server=item.ip or "-", reason="Missing server name"))
            continue
        if item.ssh_port <= 0 or item.ssh_port > 65535:
            skipped += 1
            errors.append(ServerBulkImportError(row=item.row, server=item.name, reason="Invalid SSH port"))
            continue
        if item.ip in existing_ips:
            skipped += 1
            errors.append(ServerBulkImportError(row=item.row, server=item.name, reason="IP already exists"))
            continue

        try:
            # Построение `ServerCreate` — внутри try: провайдер валидируется
            # схемой (`_checked_provider`), и `ValidationError` невалидной
            # строки обязана скипнуть только её, а не уронить весь импорт.
            payload = ServerCreate(
                name=item.name,
                ip_address=item.ip,
                ssh_user=item.ssh_user,
                ssh_port=item.ssh_port,
                provider=item.provider,
            )
            await server_service.create(db, payload, user_id)
            existing_ips.add(item.ip)
            created += 1
        except ValidationError as exc:
            # Отдельный except перед общим: голый `str(ValidationError)` — это
            # многострочный дамп pydantic со ссылкой на docs.pydantic.dev и
            # `input_value`, который целиком уезжает в `errors`/errors CSV.
            # Строка вида «поле: причина» читается, а не пугает.
            skipped += 1
            first = exc.errors()[0]
            field = ".".join(str(part) for part in first["loc"])
            errors.append(
                ServerBulkImportError(row=item.row, server=item.name, reason=f"{field}: {first['msg']}")
            )
        except Exception as exc:
            skipped += 1
            errors.append(ServerBulkImportError(row=item.row, server=item.name, reason=f"{type(exc).__name__}: {exc}"))

    csv_url = None
    if errors:
        token = store_errors_csv(build_server_errors_csv(errors))
        csv_url = f"/servers/bulk-import-errors/{token}"

    return created, skipped, errors, csv_url
